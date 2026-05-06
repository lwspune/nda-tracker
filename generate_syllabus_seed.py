"""
Parses (APJ)_2Y_25_27_Syllabus_Tracker.xlsx and writes src/lib/syllabusSeed.js
Run from project root: python generate_syllabus_seed.py
"""
import openpyxl
import json
import re
import os

XLSX_PATH = r'C:\Users\vilas\Downloads\(APJ)_2Y_25_27_Syllabus_Tracker.xlsx'
OUT_PATH = os.path.join('src', 'lib', 'syllabusSeed.js')

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

_ch_counter = [0]

def ch_id():
    _ch_counter[0] += 1
    return f'ch_{_ch_counter[0]:04d}'

def is_numbered(val):
    """True if val is a chapter Sr No (float or =formula string)."""
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str) and val.strip().startswith('='):
        return True
    return False

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def make_chapter(name, group=None):
    return {'id': ch_id(), 'name': name, 'group': group}

# ── Foundation ───────────────────────────────────────────────

def parse_science_foundation():
    """Science_Foundation sheet → split into Physics / Chemistry / Biology.
    Section headers are in col A (not numbered); chapters have numeric Sr No in col A."""
    ws = wb['Science_Foundation']
    result = {'Physics': [], 'Chemistry': [], 'Biology': []}
    current = None
    for row in list(ws.iter_rows(values_only=True))[1:]:  # skip header
        sr, name = row[0], clean(row[1])
        # Section header: col A is the subject name (string, non-numbered)
        if isinstance(sr, str) and sr.strip() in result:
            current = sr.strip()
        # Also check col B for section name (in case of layout variation)
        elif name in result:
            current = name
        elif current and name and is_numbered(sr):
            result[current].append(make_chapter(name))
    return result

def parse_foundation_flat(sheet_name):
    """Maths_Foundation and English sheets → flat chapter list."""
    ws = wb[sheet_name]
    chapters = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        sr, name = row[0], clean(row[1])
        if name and is_numbered(sr):
            chapters.append(make_chapter(name))
    return chapters

# ── State Board ──────────────────────────────────────────────

def parse_state_board_with_classes(sheet_name):
    """Mathematics, Physics, Chemistry, Biology, IT → groups: Class 11 / Class 12."""
    ws = wb[sheet_name]
    chapters = []
    current_group = None
    for row in ws.iter_rows(values_only=True):
        col_a, col_b = clean(row[0]), clean(row[1])
        if col_a and 'Class 11' in col_a:
            current_group = 'Class 11'
            continue
        if col_a and 'Class 12' in col_a:
            current_group = 'Class 12'
            continue
        if col_b and is_numbered(row[0]) and col_b not in ('Chapter', 'Sr No'):
            chapters.append(make_chapter(col_b, group=current_group))
    return chapters

def parse_geography_sb():
    ws = wb['Geography_SB']
    chapters = []
    for row in list(ws.iter_rows(values_only=True))[1:]:  # skip title row
        sr, name = row[0], clean(row[1])
        if name and is_numbered(sr):
            chapters.append(make_chapter(name))
    return chapters

# ── NDA ──────────────────────────────────────────────────────
# NDA sheets: col A = None, col B = Sr No, col C = Chapter name

def parse_nda_maths():
    ws = wb['Maths']
    chapters = []
    for row in list(ws.iter_rows(values_only=True))[2:]:  # skip title + header
        sr, name = row[1], clean(row[2])
        if name and is_numbered(sr):
            chapters.append(make_chapter(name))
    return chapters

def parse_nda_english():
    """English_NDA: section header rows (no Sr No) become groups."""
    ws = wb['English_NDA']
    chapters = []
    current_group = None
    skip_vals = {'Sr No', 'Chapter', 'Classwork', 'Lectures', 'Homework', 'Quiz', 'PYQs'}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        sr, name = row[1], clean(row[2])
        if name is None or name in skip_vals:
            continue
        if is_numbered(sr):
            chapters.append(make_chapter(name, group=current_group))
        elif sr is None and name:
            current_group = name
    return chapters

def parse_nda_science():
    """Science (NDA) → Physics / Chemistry / Biology.
    Col A = None, col B = Sr No or section header, col C = chapter name."""
    ws = wb['Science']
    result = {'Physics': [], 'Chemistry': [], 'Biology': []}
    current = None
    for row in list(ws.iter_rows(values_only=True))[1:]:
        sr = row[1]
        name = clean(row[2]) if len(row) > 2 else None
        # Section header: col B is a string matching a subject name
        if isinstance(sr, str) and sr.strip() in result:
            current = sr.strip()
        elif name in result:
            # Fallback: subject name in col C
            current = name
        elif current and name and is_numbered(sr):
            result[current].append(make_chapter(name))
    return result

def parse_nda_geo_or_polity(sheet_name):
    """Geography_NDA / Polity: numbered rows = group headers, non-numbered = chapters."""
    ws = wb[sheet_name]
    chapters = []
    current_group = None
    skip_vals = {'Sr No', 'Topic', 'Lectures', 'Homework', 'Quiz', 'PYQs', 'Sign-off'}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        sr, name = row[1], clean(row[2])
        if name is None or name in skip_vals:
            continue
        if is_numbered(sr):
            current_group = name  # group header
        elif sr is None and name:
            chapters.append(make_chapter(name, group=current_group))
    return chapters

def parse_nda_history():
    """Histry sheet: col B is Sr No for chapters, or section-header string (col C = None).
    Bullet-point sub-details under Ancient History chapters are skipped."""
    ws = wb['Histry']
    chapters = []
    current_group = None
    for row in list(ws.iter_rows(values_only=True))[1:]:
        sr, name = row[1], clean(row[2])
        # Section header: sr is a non-numbered string and name is None
        if not is_numbered(sr) and isinstance(sr, str) and name is None:
            group_name = clean(sr)
            if group_name:
                current_group = group_name
            continue
        if name is None:
            continue
        # Skip bullet-point sub-detail lines
        if name.startswith('•') or name.startswith('-') or name.startswith('−'):
            continue
        if is_numbered(sr):
            chapters.append(make_chapter(name, group=current_group))
    return chapters

def parse_nda_economics():
    """Economics: numbered rows = chapters, non-numbered sub-details skipped."""
    ws = wb['Economics']
    chapters = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        sr, name = row[1], clean(row[2])
        if name and is_numbered(sr):
            chapters.append(make_chapter(name))
    return chapters

# ── Assemble ─────────────────────────────────────────────────

def make_subj(subj_id, name, chapters):
    return {'id': subj_id, 'name': name, 'chapters': chapters}

science_f = parse_science_foundation()
science_nda = parse_nda_science()

foundation_subjects = [
    make_subj('subj_f_maths',    'Maths',     parse_foundation_flat('Maths_Foundation')),
    make_subj('subj_f_physics',  'Physics',   science_f['Physics']),
    make_subj('subj_f_chemistry','Chemistry', science_f['Chemistry']),
    make_subj('subj_f_biology',  'Biology',   science_f['Biology']),
    make_subj('subj_f_english',  'English',   parse_foundation_flat('English')),
]

state_board_subjects = [
    make_subj('subj_sb_maths',    'Mathematics', parse_state_board_with_classes('Mathematics')),
    make_subj('subj_sb_physics',  'Physics',     parse_state_board_with_classes('Physics')),
    make_subj('subj_sb_chemistry','Chemistry',   parse_state_board_with_classes('Chemistry')),
    make_subj('subj_sb_biology',  'Biology',     parse_state_board_with_classes('Biology')),
    make_subj('subj_sb_geography','Geography',   parse_geography_sb()),
    make_subj('subj_sb_it',       'IT',          parse_state_board_with_classes('IT')),
]

nda_subjects = [
    make_subj('subj_nda_maths',    'Maths',      parse_nda_maths()),
    make_subj('subj_nda_english',  'English',    parse_nda_english()),
    make_subj('subj_nda_physics',  'Physics',    science_nda['Physics']),
    make_subj('subj_nda_chemistry','Chemistry',  science_nda['Chemistry']),
    make_subj('subj_nda_biology',  'Biology',    science_nda['Biology']),
    make_subj('subj_nda_geography','Geography',  parse_nda_geo_or_polity('Geography_NDA')),
    make_subj('subj_nda_polity',   'Polity',     parse_nda_geo_or_polity('Polity')),
    make_subj('subj_nda_history',  'History',    parse_nda_history()),
    make_subj('subj_nda_economics','Economics',  parse_nda_economics()),
]

programs = [
    {
        'id': 'prog_foundation',
        'name': 'Foundation Program',
        'trackingColumns': ['Lectures', 'Worksheet', 'Quiz', 'PYQs'],
        'subjects': foundation_subjects,
    },
    {
        'id': 'prog_state_board',
        'name': 'State Board Program',
        'trackingColumns': ['Classwork', 'Homework', 'Quiz', "PYQ's", 'Sign-Off'],
        'subjects': state_board_subjects,
    },
    {
        'id': 'prog_nda',
        'name': 'NDA Program',
        'trackingColumns': ['Lectures', 'Homework', 'Quiz', 'PYQs', 'Sign-off'],
        'subjects': nda_subjects,
    },
]

# ── Stats ─────────────────────────────────────────────────────
for prog in programs:
    print(f"\n{prog['name']}:")
    for subj in prog['subjects']:
        groups = set(c['group'] for c in subj['chapters'] if c['group'])
        group_info = f"  groups: {sorted(groups)}" if groups else ''
        print(f"  {subj['name']}: {len(subj['chapters'])} chapters{group_info}")

# ── Write JS ──────────────────────────────────────────────────
js = (
    '// Auto-generated from (APJ)_2Y_25_27_Syllabus_Tracker.xlsx\n'
    '// Regenerate: python generate_syllabus_seed.py\n\n'
    f'export const SYLLABUS_SEED = {json.dumps(programs, indent=2, ensure_ascii=False)};\n'
)

os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
with open(OUT_PATH, 'w', encoding='utf-8') as f:
    f.write(js)

print(f'\nWritten to {OUT_PATH}')
print(f'Total chapters: {_ch_counter[0]}')
